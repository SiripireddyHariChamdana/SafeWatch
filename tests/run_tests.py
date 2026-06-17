"""
SafeWatch E2E Test Runner & Excel Report Generator
Discovers all Selenium and Appium E2E tests, runs them in the selected mode, 
and compiles a unified, styled Excel spreadsheet (.xlsx) test report.
"""
import os
import sys
import time
import argparse
import datetime
import unittest

# =====================================================================
# SELF-HEALING DEPENDENCY CHECK
# =====================================================================
def ensure_dependencies():
    """Ensure required libraries are installed in the environment"""
    required_libs = {
        "openpyxl": "openpyxl",
        "selenium": "selenium",
        "appium": "Appium-Python-Client"
    }
    missing_packages = []
    for module_name, pip_name in required_libs.items():
        try:
            __import__(module_name)
        except ImportError:
            missing_packages.append(pip_name)
            
    if missing_packages:
        print(f"[*] Missing dependencies found: {missing_packages}. Installing...")
        try:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing_packages)
            print("[+] Dependencies installed successfully.")
        except Exception as e:
            print(f"[!] Warning: Failed to auto-install dependencies: {e}")
            print("[!] Please install manually: pip install " + " ".join(missing_packages))

# Ensure packages before running logic
ensure_dependencies()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# CUSTOM UNITTEST TESTRESULT COLLECTOR
# =====================================================================
class SafeWatchTestResult(unittest.TestResult):
    """Custom TestResult to gather metadata from executed E2E test cases"""
    def __init__(self):
        super().__init__()
        self.results = []
        self._test_start_time = 0

    def startTest(self, test):
        super().startTest(test)
        self._test_start_time = time.time()
        print(f"[*] Running: {test.id().split('.')[-1]} ...")

    def addSuccess(self, test):
        super().addSuccess(test)
        elapsed = time.time() - self._test_start_time
        self._record_result(test, "PASS", elapsed)
        print(f"    [+] PASS ({elapsed:.3f}s)")

    def addFailure(self, test, err):
        super().addFailure(test, err)
        elapsed = time.time() - self._test_start_time
        error_msg = self._exc_info_to_string(err, test)
        self._record_result(test, "FAIL", elapsed, error_msg)
        print(f"    [-] FAIL ({elapsed:.3f}s)")

    def addError(self, test, err):
        super().addError(test, err)
        elapsed = time.time() - self._test_start_time
        error_msg = self._exc_info_to_string(err, test)
        self._record_result(test, "FAIL", elapsed, f"ERROR: {error_msg}")
        print(f"    [-] ERROR ({elapsed:.3f}s)")

    def _record_result(self, test, status, elapsed, error_details=None):
        # Extract metadata from test docstring
        doc = test._testMethodDoc or ""
        category = "General"
        description = "No description provided."
        steps = "Not specified."
        expected = "Not specified."

        # Basic parser for structured docstrings
        lines = [line.strip() for line in doc.split("\n") if line.strip()]
        desc_lines = []
        step_lines = []
        exp_lines = []
        mode = "desc"

        for line in lines:
            if line.startswith("[Category:"):
                category = line.replace("[Category:", "").replace("]", "").strip()
                continue
            elif line.startswith("Description:"):
                mode = "desc"
                desc_lines.append(line.replace("Description:", "").strip())
                continue
            elif line.startswith("Steps:"):
                mode = "steps"
                continue
            elif line.startswith("Expected:"):
                mode = "expected"
                exp_lines.append(line.replace("Expected:", "").strip())
                continue
            
            # Append based on mode
            if mode == "desc":
                desc_lines.append(line)
            elif mode == "steps":
                step_lines.append(line)
            elif mode == "expected":
                exp_lines.append(line)

        description = " ".join(desc_lines) if desc_lines else "E2E verification check."
        steps = "\n".join(step_lines) if step_lines else "Execute test automation inputs."
        expected = " ".join(exp_lines) if exp_lines else "Verification assertions succeed."

        # Identify platform
        class_name = test.__class__.__name__
        platform = "Selenium Web" if "Web" in class_name else "Appium Mobile"

        self.results.append({
            "name": test._testMethodName,
            "class": class_name,
            "platform": platform,
            "category": category,
            "description": description,
            "steps": steps,
            "expected": expected,
            "status": status,
            "time": elapsed,
            "error": error_details or "All assertions passed successfully."
        })

# =====================================================================
# REPORT GENERATION FUNCTION
# =====================================================================
def generate_excel_report(results, mode):
    """Generates a professional HSL-based spreadsheet report"""
    wb = openpyxl.Workbook()
    
    # -----------------------------------------------------------------
    # DASHBOARD SHEET
    # -----------------------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Dashboard Summary"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Header styling (navy / glassmorphism theme)
    navy_fill = PatternFill(start_color="1E2A38", end_color="1E2A38", fill_type="solid")
    accent_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    white_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="1E2A38")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    # Set titles
    ws_dash.merge_cells("A1:D1")
    ws_dash["A1"] = "SAFEWATCH SECURITY SUITE - TEST INTEGRATION DASHBOARD"
    ws_dash["A1"].font = white_font
    ws_dash["A1"].fill = navy_fill
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 40
    
    # Statistics
    total_tests = len(results)
    passed_tests = sum(1 for r in results if r["status"] == "PASS")
    failed_tests = total_tests - passed_tests
    success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0
    
    selenium_count = sum(1 for r in results if r["platform"] == "Selenium Web")
    appium_count = sum(1 for r in results if r["platform"] == "Appium Mobile")
    
    metrics = [
        ("Execution Environment", "Android Emulator & Chrome Browser (Local)"),
        ("WebDriver Model", f"{mode} Mode"),
        ("Run Date & Time", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""), # spacer
        ("Total E2E Test Cases", total_tests),
        ("Passed Alarms", passed_tests),
        ("Failed Alarms", failed_tests),
        ("Success Index Rate", f"{success_rate:.1f}%"),
        ("", ""), # spacer
        ("Selenium Web Tests", selenium_count),
        ("Appium Mobile Tests", appium_count),
    ]
    
    # Populate Dashboard values
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    row_idx = 3
    for label, val in metrics:
        if not label and not val:
            row_idx += 1
            continue
        ws_dash.cell(row=row_idx, column=2, value=label).font = bold_font
        ws_dash.cell(row=row_idx, column=2).fill = accent_fill
        ws_dash.cell(row=row_idx, column=2).border = border_thin
        
        val_cell = ws_dash.cell(row=row_idx, column=3, value=val)
        val_cell.font = regular_font
        val_cell.border = border_thin
        
        # Color Success Rate
        if "Rate" in label:
            val_cell.font = Font(name="Calibri", size=11, bold=True, color="155724")
            val_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif "Passed" in label:
            val_cell.font = Font(name="Calibri", size=11, bold=True, color="155724")
        elif "Failed" in label and val > 0:
            val_cell.font = Font(name="Calibri", size=11, bold=True, color="721C24")
            val_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
        row_idx += 1
        
    ws_dash.column_dimensions['B'].width = 30
    ws_dash.column_dimensions['C'].width = 50
    
    # -----------------------------------------------------------------
    # TEST CASES SHEET
    # -----------------------------------------------------------------
    ws_cases = wb.create_sheet(title="Test Cases Details")
    ws_cases.views.sheetView[0].showGridLines = True
    
    headers = [
        "Test ID", "Platform", "Category", "Test Method", 
        "Objective Description", "Execution Steps", "Expected Result", 
        "Actual Result / Execution Logs", "Duration (s)", "Status"
    ]
    
    # Write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws_cases.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    ws_cases.row_dimensions[1].height = 28
    
    # Fills for status
    pass_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # soft green
    pass_font = Font(name="Calibri", size=11, bold=True, color="385723")
    
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft red
    fail_font = Font(name="Calibri", size=11, bold=True, color="C00000")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Populate rows
    for idx, r in enumerate(results, 1):
        row_num = idx + 1
        ws_cases.row_dimensions[row_num].height = 55  # readable height
        
        tc_id = f"TC-{idx:03d}"
        
        c1 = ws_cases.cell(row=row_num, column=1, value=tc_id)
        c1.alignment = center_align
        c1.font = bold_font
        
        c2 = ws_cases.cell(row=row_num, column=2, value=r["platform"])
        c2.alignment = center_align
        
        c3 = ws_cases.cell(row=row_num, column=3, value=r["category"])
        c3.alignment = center_align
        
        c4 = ws_cases.cell(row=row_num, column=4, value=r["name"])
        c4.alignment = left_wrap_align
        
        c5 = ws_cases.cell(row=row_num, column=5, value=r["description"])
        c5.alignment = left_wrap_align
        
        c6 = ws_cases.cell(row=row_num, column=6, value=r["steps"])
        c6.alignment = left_wrap_align
        
        c7 = ws_cases.cell(row=row_num, column=7, value=r["expected"])
        c7.alignment = left_wrap_align
        
        c8 = ws_cases.cell(row=row_num, column=8, value=r["error"])
        c8.alignment = left_wrap_align
        
        c9 = ws_cases.cell(row=row_num, column=9, value=round(r["time"], 4))
        c9.alignment = center_align
        
        c10 = ws_cases.cell(row=row_num, column=10, value=r["status"])
        c10.alignment = center_align
        
        # Color formatting by status
        if r["status"] == "PASS":
            c10.fill = pass_fill
            c10.font = pass_font
        else:
            c10.fill = fail_fill
            c10.font = fail_font
            
        # Draw borders
        for col in range(1, 11):
            ws_cases.cell(row=row_num, column=col).border = border_thin
            if col != 10 and col != 1:
                ws_cases.cell(row=row_num, column=col).font = regular_font
                
    # Auto-adjust column widths with safety bounds
    for col in ws_cases.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Find maximum length, check up to first 15 rows to prevent massive column expansion
        for cell in col[:15]:
            if cell.value:
                # split lines to measure wrap size
                lines = str(cell.value).split('\n')
                for l in lines:
                    if len(l) > max_len:
                        max_len = len(l)
                        
        ws_cases.column_dimensions[col_letter].width = max(12, min(45, max_len + 3))
        
    # Specific spacing override for descriptive columns
    ws_cases.column_dimensions['E'].width = 30 # Description
    ws_cases.column_dimensions['F'].width = 30 # Steps
    ws_cases.column_dimensions['G'].width = 30 # Expected
    ws_cases.column_dimensions['H'].width = 30 # Actual Logs
    
    # Save file to root workspace
    report_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "safewatch_test_report.xlsx"))
    wb.save(report_path)
    return report_path

# =====================================================================
# MAIN RUNNER PROTOCOL
# =====================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run SafeWatch E2E Test Suite")
    parser.add_argument("--real", action="store_true", help="Execute using real WebDrivers")
    args = parser.parse_args()
    
    mode = "REAL" if args.real else "SIMULATED"
    os.environ["SAFEWATCH_TEST_MODE"] = mode
    
    print("=====================================================================")
    print(f"        SAFEWATCH E2E TEST RUNNER - {mode} MODE")
    print("=====================================================================")
    print(f"[*] Discovery Path: {os.path.dirname(__file__)}")
    print("[*] Gathering 100 test cases for Appium and Selenium...")
    
    # Locate files explicitly or via loader
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    tests_dir = os.path.abspath(os.path.dirname(__file__))
    
    # Discover Selenium Web Tests
    selenium_dir = os.path.join(tests_dir, "selenium")
    selenium_suite = loader.discover(start_dir=selenium_dir, pattern="test_*.py", top_level_dir=tests_dir)
    suite.addTests(selenium_suite)
    
    # Discover Appium Mobile Tests
    appium_dir = os.path.join(tests_dir, "appium")
    appium_suite = loader.discover(start_dir=appium_dir, pattern="test_*.py", top_level_dir=tests_dir)
    suite.addTests(appium_suite)
    
    total_found = suite.countTestCases()
    print(f"[+] Successfully loaded {total_found} test cases.")
    
    if total_found == 0:
        print("[!] Error: No test cases found! Verify directories and filenames.")
        sys.exit(1)
        
    print("[*] Initiating test run sequence...")
    runner_result = SafeWatchTestResult()
    
    start_time = time.time()
    suite.run(runner_result)
    end_time = time.time()
    
    print("\n=====================================================================")
    print("        TEST SUITE SEQUENCE EXECUTION COMPLETE")
    print("=====================================================================")
    print(f"[+] Total Duration: {end_time - start_time:.2f} seconds")
    print(f"[+] Passed: {len(runner_result.results) - len(runner_result.failures) - len(runner_result.errors)} / {total_found}")
    print(f"[+] Failed/Errors: {len(runner_result.failures) + len(runner_result.errors)} / {total_found}")
    
    print("\n[*] Compiling results and generating Excel analysis report...")
    try:
        report_file = generate_excel_report(runner_result.results, mode)
        print(f"[+] Excel report created at: {report_file}")
    except Exception as e:
        print(f"[!] Failed to generate Excel report: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(2)
        
    print("=====================================================================")
